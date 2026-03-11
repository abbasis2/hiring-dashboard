from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

OUTSTANDING_SHEET = "Outstanding Roles"
FILLED_SHEET = "Filled Roles"

OUTSTANDING_HEADER_ROW = 2
FILLED_HEADER_ROW = 2


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.date().isoformat()
        except Exception:
            return value.isoformat()
    return value


def sheet_rows(workbook, sheet_name: str, header_row: int) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    headers = [cell for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rows.append({header: normalize_cell(value) for header, value in zip(headers, row)})
    return rows


def parse_excel_workbook(file_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    if OUTSTANDING_SHEET not in workbook.sheetnames or FILLED_SHEET not in workbook.sheetnames:
        raise ValueError("Workbook must include 'Outstanding Roles' and 'Filled Roles' tabs")
    return {
        "outstanding_roles": sheet_rows(workbook, OUTSTANDING_SHEET, OUTSTANDING_HEADER_ROW),
        "filled_roles": sheet_rows(workbook, FILLED_SHEET, FILLED_HEADER_ROW),
    }
